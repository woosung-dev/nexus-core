# 레드팀 피드백 xlsx 를 추출·정규화하고 집계 JSON 을 만든다
import json
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import load_workbook

SRC = "/Users/woosung/Downloads/축복·가정관리 AI 상담 챗봇 테스트 및 피드백 (레드팀)(응답).xlsx"
OUT_DATA = "/Users/woosung/project/agy-project/nexus-core/exports/_redteam_data.json"


def norm_user(v):
    v = (v or "").strip()
    return {"이보": "이보영", "김관": "김관우"}.get(v, v)


def norm_qtype(v):
    v = (v or "").strip()
    if v.startswith("축복 준비 및 매칭"):
        return "축복 준비·매칭"
    if v.startswith("축복유형"):
        return "축복 유형(1세/2세/은사 등)"
    if v.startswith("축복정리"):
        return "축복 정리"
    if v.startswith("가정출발"):
        return "가정 출발"
    if "탈선" in v:
        return "탈선·성적 문제"
    return v or "(미분류)"


wb = load_workbook(SRC, data_only=True)
ws = wb.active
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]

records = []
for r in rows:
    ts = r[0]
    score = None
    try:
        score = float(r[5]) if r[5] not in (None, "") else None
    except (ValueError, TypeError):
        score = None
    areas = []
    if r[6]:
        areas = [p.strip() for p in str(r[6]).split(",") if p.strip()]
    records.append({
        "ts": ts.isoformat() if isinstance(ts, datetime) else str(ts),
        "date": ts.date().isoformat() if isinstance(ts, datetime) else None,
        "user": norm_user(r[1]),
        "qtype": norm_qtype(r[2]),
        "question": (r[3] or "").strip(),
        "answer": (r[4] or "").strip(),
        "score": score,
        "areas": areas,
        "feedback": (r[7] or "").strip(),
        "keywords": (r[8] or "").strip(),
        "etc": (r[9] or "").strip(),
    })

# 집계
qtype_cnt = Counter(x["qtype"] for x in records)
user_cnt = Counter(x["user"] for x in records)
score_cnt = Counter(int(x["score"]) for x in records if x["score"] is not None)
area_cnt = Counter()
for x in records:
    for a in x["areas"]:
        area_cnt[a] += 1
date_cnt = Counter(x["date"] for x in records if x["date"])

scores = [x["score"] for x in records if x["score"] is not None]
avg = round(sum(scores) / len(scores), 2) if scores else 0
neg = sum(1 for s in scores if s <= 2)
pos = sum(1 for s in scores if s >= 4)

# 질문유형별 평균점수
qtype_scores = defaultdict(list)
for x in records:
    if x["score"] is not None:
        qtype_scores[x["qtype"]].append(x["score"])
qtype_avg = {k: round(sum(v) / len(v), 2) for k, v in qtype_scores.items()}

agg = {
    "total": len(records),
    "users": len(user_cnt),
    "rated": len(scores),
    "avg_score": avg,
    "neg_count": neg, "neg_pct": round(neg / len(scores) * 100) if scores else 0,
    "pos_count": pos, "pos_pct": round(pos / len(scores) * 100) if scores else 0,
    "qtype": dict(qtype_cnt),
    "user": dict(user_cnt),
    "score": {str(k): score_cnt.get(k, 0) for k in range(1, 6)},
    "area": dict(area_cnt.most_common()),
    "date": dict(sorted(date_cnt.items())),
    "qtype_avg": qtype_avg,
    "ts_min": min(x["ts"] for x in records),
    "ts_max": max(x["ts"] for x in records),
}

json.dump({"records": records, "agg": agg}, open(OUT_DATA, "w"), ensure_ascii=False, indent=1)
print("저장:", OUT_DATA)
print(json.dumps(agg, ensure_ascii=False, indent=1))
