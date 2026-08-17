# Neon 실서버에서 최근 7일 대화 전문을 읽어 .xlsx 로 정리하는 일회성 스크립트 (읽기 전용)
import asyncio
import datetime as dt
from zoneinfo import ZoneInfo

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

URL = "postgresql://neondb_owner:npg_l8Xgu2JIxnVA@ep-icy-wave-amjjo0k9-pooler.c-5.us-east-1.aws.neon.tech/neondb?sslmode=require"
KST = ZoneInfo("Asia/Seoul")
EXCLUDE_EMAIL = "woosung@test.com"
START = "2026-05-21"   # 포함
END = "2026-05-29"     # 미포함 (05-28 까지)
OUT = "/Users/woosung/project/agy-project/nexus-core/exports/nexus_chats_2026-05-21_to_05-28.xlsx"

ROLE_KO = {"USER": "사용자", "ASSISTANT": "AI", "user": "사용자", "assistant": "AI"}
FB_KO = {"LIKE": "좋아요", "DISLIKE": "싫어요", "like": "좋아요", "dislike": "싫어요"}


def kst(ts):
    if ts is None:
        return ""
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=dt.timezone.utc)
    return ts.astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")


async def fetch():
    conn = await asyncpg.connect(URL)
    try:
        rows = await conn.fetch(
            """
            SELECT s.id AS session_id, b.name AS bot_name, u.email AS user_email,
                   s.title AS session_title, s.created_at AS session_created,
                   m.id AS message_id, m.role::text AS role, m.content AS content,
                   m.feedback AS feedback, m.feedback_reasons AS fb_reasons,
                   m.feedback_comment AS fb_comment, m.created_at AS msg_created
            FROM chat_sessions s
            JOIN users u ON u.id = s.user_id
            LEFT JOIN bots b ON b.id = s.bot_id
            LEFT JOIN messages m ON m.session_id = s.id
            WHERE u.email <> $1
              AND s.created_at >= $2
              AND s.created_at <  $3
            ORDER BY s.created_at ASC, s.id ASC, m.created_at ASC, m.id ASC
            """,
            EXCLUDE_EMAIL, dt.date.fromisoformat(START), dt.date.fromisoformat(END),
        )
        return rows
    finally:
        await conn.close()


def build(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "대화 전문"

    headers = [
        "세션ID", "봇", "사용자", "세션 제목", "세션 생성(KST)",
        "메시지ID", "역할", "내용", "피드백", "피드백 사유", "피드백 코멘트", "메시지 시각(KST)",
    ]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="2F4858")
    head_font = Font(bold=True, color="FFFFFF")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    user_fill = PatternFill("solid", fgColor="EAF2F8")
    sess_count = set()
    msg_count = 0
    for r in rows:
        if r["message_id"] is None:
            continue  # 메시지 없는 빈 세션은 스킵
        sess_count.add(r["session_id"])
        msg_count += 1
        ws.append([
            r["session_id"], r["bot_name"] or "", r["user_email"],
            r["session_title"] or "", kst(r["session_created"]),
            r["message_id"], ROLE_KO.get(r["role"], r["role"]),
            r["content"] or "", FB_KO.get(r["feedback"], r["feedback"] or ""),
            r["fb_reasons"] or "", r["fb_comment"] or "", kst(r["msg_created"]),
        ])
        if r["role"] in ("USER", "user"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).fill = user_fill

    # 컬럼 너비/정렬
    widths = [8, 14, 22, 30, 18, 9, 7, 70, 9, 18, 24, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wrap_cols = {4, 8, 10, 11}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(cell.column in wrap_cols),
            )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(OUT)
    return len(sess_count), msg_count


async def main():
    rows = await fetch()
    n_sess, n_msg = build(rows)
    print(f"저장 완료: {OUT}")
    print(f"세션 {n_sess}개 / 메시지 {n_msg}건")


asyncio.run(main())
