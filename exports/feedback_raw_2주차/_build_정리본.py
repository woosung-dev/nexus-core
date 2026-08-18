# 2주차 레드팀 응답(인간 원문) → 평가자·회차·질문·피드백 원문 그대로 + 영역 태그만 부착한 정리본 생성
"""
판단(채택/기각) 없음. 원문 보존 + 영역 태그(프롬프트/RAG/코드/정책/미분류, 복수 허용)만 부착.
태그는 codex CLT 분류(과금 0). --no-tag 면 태그 없이 원문 시트만 생성(읽기/쓰기 검증용).
출력: exports/feedback_raw_2주차/2주차_피드백_원문_정리본.xlsx + 요약.md
"""

import argparse
import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path("/Users/woosung/Downloads/축복·가정관리 AI 상담 챗봇 테스트 및 피드백 v2주차(레드팀)(응답) (1).xlsx")
OUT_DIR = Path(__file__).resolve().parent
OUT_XLSX = OUT_DIR / "2주차_피드백_원문_정리본.xlsx"
OUT_MD = OUT_DIR / "요약.md"
REPO = Path("/Users/woosung/project/agy-project/nexus-core")

# 출력 컬럼 (원문 그대로 — 헤더만 짧게 재명명, 데이터는 무변형)
COLS = [
    ("행", None),
    ("회차(타임스탬프)", 0),
    ("평가자", 1),
    ("질문유형", 2),
    ("질문원문", 3),
    ("응답 A(통합)", 4),
    ("응답 B(원리)", 5),
    ("응답 C(정밀)", 6),
    ("베스트봇 선택", 7),
    ("피드백 원문(좋은점/아쉬운점)", 8),
    ("기타의견·버그", 9),
]

TAG_INSTRUCTION = (
    "너는 챗봇 사용자 피드백을 '영역'으로 분류하는 분류기다. 채택/기각이나 우선순위 판단은 절대 하지 마라.\n"
    "<stdin> JSON 배열의 각 항목에는 idx, 피드백, 기타의견이 있다. 그 내용이 어떤 영역에 대한 지적인지 태그하라.\n"
    "영역 정의:\n"
    "- 프롬프트: 응답의 내용·톤·말투·공감·길이·구성·답변 방식 등 생성 방식에 대한 지적\n"
    "- RAG: 사실 정확성·출처·규정/문서 내용·정보 누락·최신성 등 지식/근거에 대한 지적\n"
    "- 코드: 버그·에러·속도·로딩·UI·화면·기능 오작동 등 시스템에 대한 지적\n"
    "- 정책: 위기·민감 대응·안전·낙인·상담 연계·공개 수위·윤리 등 정책 판단에 대한 지적\n"
    "한 항목이 여러 영역에 걸치면 복수 태그를 붙여라. 어느 영역인지 명확하지 않으면 ['미분류'].\n"
    "설명 없이 입력과 같은 개수·순서의 JSON 배열만 출력하라. 각 원소: {\"idx\": 정수, \"tags\": [문자열...]}."
)
VALID_TAGS = {"프롬프트", "RAG", "코드", "정책", "미분류"}


def read_source():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [r for r in ws.iter_rows(values_only=True)][1:]
    wb.close()
    data = [r for r in rows if r and r[0]]  # 타임스탬프 있는 행만
    return data


def _extract_json_array(text):
    t = re.sub(r"^```(?:json)?", "", text.strip()).strip()
    t = re.sub(r"```$", "", t).strip()
    i, j = t.find("["), t.rfind("]")
    return json.loads(t[i : j + 1])


def codex_tag(items):
    p = subprocess.run(
        ["codex", "exec", TAG_INSTRUCTION, "-s", "read-only", "-c", 'model_reasoning_effort="low"'],
        input=json.dumps(items, ensure_ascii=False),
        capture_output=True, text=True, cwd=str(REPO), timeout=600,
    )
    if p.returncode != 0:
        raise RuntimeError(f"codex exit {p.returncode}: {p.stderr[-300:]}")
    return _extract_json_array(p.stdout)


def tag_all(data, batch=25):
    """행별 영역 태그 리스트 반환. codex 실패/누락 행은 미분류로 보존(데이터 손실 방지)."""
    tags_by_idx = {}
    for start in range(0, len(data), batch):
        chunk = data[start : start + batch]
        items = [
            {
                "idx": start + k,
                "피드백": (r[8] if len(r) > 8 and r[8] else ""),
                "기타의견": (r[9] if len(r) > 9 and r[9] else ""),
            }
            for k, r in enumerate(chunk)
        ]
        print(f"  codex 태그 {start}~{start + len(chunk) - 1} …", flush=True)
        try:
            res = codex_tag(items)
            for e in res:
                idx = e.get("idx")
                tags = [t for t in (e.get("tags") or []) if t in VALID_TAGS]
                tags_by_idx[idx] = tags or ["미분류"]
        except Exception as e:
            print(f"    배치 실패({e}) → 미분류 보존", flush=True)
    return [tags_by_idx.get(i, ["미분류"]) for i in range(len(data))]


def write_xlsx(data, tags):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2주차 인간 원문+태그"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="375623")
    headers = [c[0] for c in COLS] + (["영역태그(자동·검토보조)"] if tags is not None else [])
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for i, r in enumerate(data):
        out = []
        for name, src in COLS:
            if src is None:
                out.append(i + 1)
            else:
                out.append(r[src] if src < len(r) else None)  # 뒤쪽 빈 셀이 생략된 행 방어
        if tags is not None:
            out.append(", ".join(tags[i]))
        ws.append(out)

    widths = [5, 20, 10, 16, 34, 30, 30, 30, 10, 50, 28, 16]
    for ci, w in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_md(data, tags):
    who = Counter(str(r[1]).strip() for r in data if r[1])
    dates = Counter(str(r[0])[:10] for r in data if r[0])
    lines = [
        "# 2주차 레드팀 피드백 — 인간 원문 정리본 (판단 없음)",
        "",
        f"- 원본: `{SRC.name}` (Google Forms 응답)",
        f"- 데이터 행수: {len(data)}",
        "- 성격: **2주차 인간 원문**. AI 페르소나 평가(3회차)와 혼용 금지 (크로스체크 원칙).",
        "- 영역 태그는 codex 자동 분류(검토 보조) — 채택/기각 판단 아님. 애매하면 '미분류'.",
        "",
        "## 평가자별 제출 수",
        "",
    ]
    for k, v in who.most_common():
        lines.append(f"- {k}: {v}")
    lines += ["", "## 회차(날짜)별 제출 수", ""]
    for k, v in sorted(dates.items()):
        lines.append(f"- {k}: {v}")
    if tags is not None:
        tag_count = Counter(t for ts in tags for t in ts)
        lines += ["", "## 영역 태그 분포 (복수 태그 포함)", ""]
        for k, v in tag_count.most_common():
            lines.append(f"- {k}: {v}")
    lines += [
        "",
        "## 주의",
        "- 평가자명은 원문 그대로 보존했다('김관우'와 '김관'은 동일인 추정 오타이나 임의 수정하지 않음).",
        "- 태그는 분류 보조일 뿐이며, 모든 원문이 보존되어 있어 태그를 무시하고 재판단할 수 있다.",
    ]
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-tag", action="store_true", help="태그 없이 원문 시트만 (검증용)")
    args = ap.parse_args()

    data = read_source()
    print(f"원본 데이터 {len(data)}행 로드")
    tags = None if args.no_tag else tag_all(data)
    write_xlsx(data, tags)
    write_md(data, tags)
    print(f"저장: {OUT_XLSX.name}, {OUT_MD.name}")
    if tags is not None:
        tc = Counter(t for ts in tags for t in ts)
        print("태그 분포:", dict(tc))


if __name__ == "__main__":
    main()
