# 레드팀 v2 xlsx(A/B/C 택1 비교)를 추출·정규화하고 봇별 승수 집계 JSON 생성
import json
import re
from collections import Counter, defaultdict
from datetime import datetime

from openpyxl import load_workbook

SRC = "/Users/woosung/Downloads/축복·가정관리 AI 상담 챗봇 테스트 및 피드백 v2주차(레드팀)(응답).xlsx"
OUT = "/Users/woosung/project/agy-project/nexus-core/exports/_redteam_v2_data.json"

# A=통합, B=원리, C=정밀
CHOICE_BOT = {"A": "통합", "B": "원리", "C": "정밀"}


def norm_user(v):
    v = (v or "").strip()
    return {"김관": "김관우", "이보": "이보영"}.get(v, v)


def norm_qtype(v):
    v = (v or "").strip()
    if v.startswith("축복 준비"):
        return "축복 준비·매칭"
    if v.startswith("축복유형"):
        return "축복 유형(은사 등)"
    if v.startswith("축복정리"):
        return "축복 정리"
    if v.startswith("가정출발"):
        return "가정 출발"
    if "탈선" in v:
        return "탈선·성적 문제"
    return v or "(미분류)"


def norm_choice(v):
    """선택값을 A/B/C(단일) · 복수 · 무효 로 정규화."""
    s = str(v or "").strip()
    if not s or s == "None":
        return "무효"
    # 무응답·모름류
    if any(k in s for k in ["없음", "없", "모르겠", "응답이 없", "응답이 나오지", "듣고 싶은", "확인필요"]):
        return "무효"
    # 셋 다·동일류 → 복수
    if any(k in s for k in ["모두", "3가지", "세가지", "세 가지", "동일"]):
        return "복수"
    letters = {ch.upper() for ch in re.findall(r"[abcABC]", s)}
    if len(letters) == 1:
        return letters.pop()
    if len(letters) >= 2:
        return "복수"
    return "무효"


wb = load_workbook(SRC, data_only=True)
ws = wb.active
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(r)]

records = []
for r in rows:
    ts = r[0]
    choice = norm_choice(r[7])
    records.append({
        "ts": ts.isoformat() if isinstance(ts, datetime) else str(ts),
        "date": ts.date().isoformat() if isinstance(ts, datetime) else None,
        "user": norm_user(r[1]),
        "qtype": norm_qtype(r[2]),
        "q": (r[3] or "").strip(),
        "ansA": (r[4] or "").strip(),
        "ansB": (r[5] or "").strip(),
        "ansC": (r[6] or "").strip(),
        "choice": choice,                       # A/B/C/복수/무효
        "win": CHOICE_BOT.get(choice),          # 통합/원리/정밀 or None
        "raw_choice": str(r[7] or "").strip(),
        "feedback": (r[8] or "").strip(),
        "etc": (r[9] or "").strip(),
    })

# 집계
win = Counter(x["win"] for x in records if x["win"])
multi = sum(1 for x in records if x["choice"] == "복수")
invalid = sum(1 for x in records if x["choice"] == "무효")
valid_total = sum(win.values())

win_by_qtype = defaultdict(lambda: Counter())
win_by_tester = defaultdict(lambda: Counter())
for x in records:
    if x["win"]:
        win_by_qtype[x["qtype"]][x["win"]] += 1
        win_by_tester[x["user"]][x["win"]] += 1

qtype_cnt = Counter(x["qtype"] for x in records)
user_cnt = Counter(x["user"] for x in records)
date_cnt = Counter(x["date"] for x in records if x["date"])

BOTS = ["통합", "원리", "정밀"]
agg = {
    "total": len(records),
    "testers": [u for u, _ in user_cnt.most_common()],
    "qtypes": [q for q, _ in qtype_cnt.most_common()],
    "bots": BOTS,
    "win": {b: win.get(b, 0) for b in BOTS},
    "win_pct": {b: round(win.get(b, 0) / valid_total * 100) if valid_total else 0 for b in BOTS},
    "valid_total": valid_total,
    "multi": multi,
    "invalid": invalid,
    "win_by_qtype": {q: {b: win_by_qtype[q].get(b, 0) for b in BOTS} for q in qtype_cnt},
    "win_by_tester": {u: {b: win_by_tester[u].get(b, 0) for b in BOTS} for u in user_cnt},
    "qtype_cnt": dict(qtype_cnt),
    "user_cnt": dict(user_cnt),
    "date": dict(sorted(date_cnt.items())),
    "ts_min": min(x["ts"] for x in records),
    "ts_max": max(x["ts"] for x in records),
}

json.dump({"records": records, "agg": agg}, open(OUT, "w"), ensure_ascii=False, indent=1)
print("저장:", OUT)
print(f"총 {agg['total']}행 / 유효택1 {valid_total} (복수 {multi}, 무효 {invalid})  합계검증={valid_total+multi+invalid}")
print("봇별 승수:", agg["win"], "승률%:", agg["win_pct"])
print("원본 선택값 분포:", dict(Counter(x["raw_choice"] for x in records).most_common()[:15]))
