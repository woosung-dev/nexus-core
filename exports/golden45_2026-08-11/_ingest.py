# 관리자 회신 정답지(xlsx 시트 ①) → exports/regression/questions.json 의 45문항에 골든 주입.
#
# 지금까지 45문항은 전부 `golden` 이 없어 `_l3.py` 채점 분모에서 빠져 있었다
# (채점된 것은 불변제약 C 10건뿐). 이 스크립트가 그 공백을 메운다.
#
# 파싱 규약은 `exports/prompt4_2026-08-05/_build45.py` 를 그대로 따른다 — NFC 정규화,
# 헤더 4행·데이터 5행~, 매칭 키는 번호(`#` ↔ `no`).
#
# 손대지 않는 것: C 불변제약 10건(자체 골든 보유) · anchors · evidence_status ·
#                must_any · must_not. 이 스크립트는 **덧붙이기만** 한다.
#
# DB 를 쓰지 않는다. 라이브 `redteam_goldens` 적재는 채점 결과를 보고 별도로 판단한다.
import argparse
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path("/Users/woosung/project/agy-project/nexus-core")
XLSX = Path.home() / "Downloads" / "축복 챗봇 정답지 요청 0806 (1).xlsx"
SHEET = "① 45문항 정답지"
QJSON = ROOT / "exports" / "regression" / "questions.json"
BACKUP = ROOT / "exports" / "regression" / "questions_45set_pre_golden_2026-08-11.json"
DUMP = Path(__file__).parent / "goldens_45.json"

HDR_ROW = 4
FIRST_ROW = 5
COL = {"no": 2, "q": 3, "golden": 4, "source": 5, "safe_ok": 6, "risk": 7, "cat": 8}


def nfc(s):
    return unicodedata.normalize("NFC", str(s or "")).strip()


def norm(s):
    """_build45.py 와 같은 규칙 — NFC + 영숫자/한글만 남긴다."""
    return "".join(ch for ch in nfc(s) if ch.isalnum())


def read_xlsx():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    hdr = [nfc(c.value) for c in ws[HDR_ROW]]
    assert hdr[COL["golden"] - 1].startswith("①"), f"열 배치가 다르다: {hdr}"
    assert hdr[COL["source"] - 1].startswith("②"), f"열 배치가 다르다: {hdr}"
    assert hdr[COL["safe_ok"] - 1].startswith("③"), f"열 배치가 다르다: {hdr}"

    out = {}
    for row in ws.iter_rows(min_row=FIRST_ROW, max_row=ws.max_row):
        raw = row[COL["no"] - 1].value
        if raw is None:
            continue
        no = int(raw)
        g = nfc(row[COL["golden"] - 1].value)
        if not g:
            raise SystemExit(f"#{no} 정답 칸이 비었다 — 회신본이 맞는지 확인하라")
        src = nfc(row[COL["source"] - 1].value)
        safe = nfc(row[COL["safe_ok"] - 1].value)
        out[no] = {
            "no": no,
            "q": nfc(row[COL["q"] - 1].value),
            "golden": re.sub(r"[ \t]+", " ", g),
            "golden_source": src or None,
            # 「예」/「아니오」 외의 표기가 오면 판단을 미룬다 — 조용히 true 로 만들지 않는다.
            "safe_ok": True if safe.startswith("예") else (False if safe else None),
            "safe_ok_raw": safe or None,
        }
    assert len(out) == 45, f"45행이 아니다: {len(out)}"
    return out


def main(apply):
    gold = read_xlsx()
    doc = json.loads(QJSON.read_text(encoding="utf-8"))
    items = doc["items"]

    matched, mismatched, skipped = 0, [], 0
    for it in items:
        no = it.get("no")
        if no is None:                       # C 불변제약 10건 — 건드리지 않는다
            skipped += 1
            continue
        g = gold.get(no)
        if g is None:
            mismatched.append((no, "xlsx 에 해당 번호 없음"))
            continue
        if norm(g["q"]) != norm(it["q"]):
            mismatched.append((no, f"질문 원문 불일치\n      xlsx: {g['q'][:60]}\n      json: {it['q'][:60]}"))
            continue
        it["golden"] = g["golden"]
        it["golden_source"] = g["golden_source"]
        it["safe_ok"] = g["safe_ok"]
        matched += 1

    doc["golden_source_file"] = XLSX.name
    doc["golden_ingested_at"] = "2026-08-11"

    have = sum(1 for i in items if i.get("golden"))
    print(f"매칭 {matched}/45 · C 건너뜀 {skipped} · golden 보유 {have}/{len(items)}")
    if mismatched:
        for no, why in mismatched:
            print(f"  ⚠ #{no} {why}")
        raise SystemExit("불일치가 있어 쓰지 않는다")

    safe_yes = sum(1 for i in items if i.get("safe_ok") is True)
    safe_no = sum(1 for i in items if i.get("safe_ok") is False)
    unclear = [i["no"] for i in items if i.get("no") and i.get("safe_ok") is None]
    src = sum(1 for i in items if i.get("golden_source"))
    print(f"안전응대 인정: 예 {safe_yes} · 아니오 {safe_no} · 불명 {len(unclear)}{unclear or ''}")
    print(f"관리자 근거 기재: {src}/45")

    DUMP.write_text(json.dumps(list(gold.values()), ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"원본 추출 보존 → {DUMP.relative_to(ROOT)}")

    if not apply:
        print("\n(dry-run) --apply 를 붙여야 questions.json 에 쓴다")
        return

    if not BACKUP.exists():
        BACKUP.write_text(QJSON.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"백업 → {BACKUP.name}")
    QJSON.write_text(json.dumps(doc, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"기록 → {QJSON.relative_to(ROOT)}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    main(ap.parse_args().apply)
