# 정답지 xlsx → 측정용 questions.json. API 호출 0.
#
# 시트 ①에서 질문·정답·근거·위험도·카테고리를, 시트 ⑤에서 채점 키워드와 기존 충족률을 뽑는다.
# 정답은 45문항 중 10건만 왔다(2026-08-06 회신). 나머지는 키워드로만 채점한다.
import json
import re
from pathlib import Path

import openpyxl

XLSX = Path.home() / "Downloads" / "축복 챗봇 정답지 요청 0806.xlsx"
OUT = Path(__file__).parent / "questions.json"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    kw: dict[int, dict] = {}
    for r in wb["⑤ 필수 키워드"].iter_rows(min_row=5, values_only=True):
        if not r or not r[1]:
            continue
        m = re.match(r"(\d+)", str(r[5] or ""))
        kw[int(r[1])] = {
            "keywords": [k.strip() for k in str(r[3] or "").split(",") if k.strip()],
            "baseline_pct": int(m.group(1)) if m else None,
        }

    rows = []
    for r in wb["① 45문항 정답지"].iter_rows(min_row=5, values_only=True):
        if not r or not r[1]:
            continue
        n = int(r[1])
        rows.append(
            {
                "n": n,
                "question": str(r[2]).strip(),
                "golden": str(r[3]).strip() if r[3] else None,
                "golden_basis": str(r[4]).strip() if r[4] else None,
                "refusal_ok": str(r[5]).strip() if r[5] else None,
                "risk": str(r[6]).strip() if r[6] else None,
                "category": str(r[7]).strip() if r[7] else None,
                "keywords": kw.get(n, {}).get("keywords", []),
                "baseline_pct": kw.get(n, {}).get("baseline_pct"),
            }
        )

    OUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    golden = sum(1 for x in rows if x["golden"])
    avg_kw = sum(len(x["keywords"]) for x in rows) / len(rows)
    print(f"{len(rows)}문항 → {OUT}")
    print(f"  정답 있는 문항 {golden}건 · 문항당 키워드 평균 {avg_kw:.1f}개")


if __name__ == "__main__":
    main()
