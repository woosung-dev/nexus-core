# 3회차 난이도'중' 18문항을 5개 봇(id5·3·21·22·16)에 돌려 응답+인용을 응답_전체.json으로 저장
import asyncio
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import psycopg2

sys.path.insert(0, "/Users/woosung/project/agy-project/nexus-core/backend")
from app.services.rag.gemini import GeminiRAGService  # noqa: E402

XLSX = Path("/Users/woosung/Downloads/축복챗봇_3회차_질문지_정답지_v1.xlsx")
ROOT = Path(__file__).resolve().parent.parent  # 이 스크립트가 속한 테스트 폴더
OUT = ROOT / "봇별질문응답" / "_data" / "응답_전체.json"
MODEL = "gemini-3.1-flash-lite"
TEMPERATURE = 0.2
MAX_TOKENS = 1500
DIFFICULTY = os.environ.get("TEST_DIFFICULTY", "중")

BOT_ORDER = [5, 3, 21, 22, 16]
BOT_META = {
    5: {"성격": "정밀기반", "rag_docs": 15},
    3: {"성격": "통합기반", "rag_docs": 15},
    21: {"성격": "정밀 full", "rag_docs": 15},
    22: {"성격": "통합 full", "rag_docs": 15},
    16: {"성격": "완전체(검증프롬프트+풀RAG)", "rag_docs": 15},
}
FILE_SLUG = {5: "블레싱가_id5", 3: "블레싱나_id3", 21: "블레싱정밀full_id21",
             22: "블레싱통합full_id22", 16: "달인봇정밀full_id16"}


def load_questions():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    q = wb["① 질문지"]
    a = wb["② 정답지 (초안)"]
    # 정답지를 ID로 인덱싱
    ans = {}
    for row in a.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        ans[row[0]] = {
            "golden_must": str(row[3] or "").strip(),     # 골든답변 — 필수 포함 요소
            "golden_avoid": str(row[4] or "").strip(),    # 금지 · 주의 요소
            "golden_routing": str(row[5] or "").strip(),  # 라우팅 요구
            "golden_severity": str(row[6] or "").strip(), # 실패 시 심각도(후보)
            "golden_basis": str(row[7] or "").strip(),    # 근거(2주차 피드백)
        }
    out = []
    for row in q.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[1] != DIFFICULTY:
            continue
        qid = row[0]
        out.append({
            "id": qid,
            "category": str(row[2] or "").strip(),
            "q": str(row[3] or "").strip(),
            "source": str(row[4] or "").strip(),
            **ans.get(qid, {}),
        })
    return out


def load_bots():
    raw = os.environ["DATABASE_URL"]
    url = raw.replace("+asyncpg", "").replace("@db:", "@localhost:")
    conn = psycopg2.connect(url)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, name, llm_model, system_prompt FROM bots WHERE id = ANY(%s)",
        (BOT_ORDER,),
    )
    rows = {r[0]: {"id": r[0], "name": r[1], "llm_model": r[2], "system_prompt": r[3]}
            for r in cur.fetchall()}
    conn.close()
    return [rows[i] for i in BOT_ORDER]


async def call(rag, bot_id, system_prompt, q, tries=5):
    delay = 20
    for i in range(tries):
        try:
            resp = await asyncio.wait_for(
                rag.generate_with_rag(bot_id=bot_id, prompt=q, system_prompt=system_prompt,
                                      model_name=MODEL, temperature=TEMPERATURE, max_tokens=MAX_TOKENS),
                timeout=80)
            return resp.answer, [c.title for c in resp.citations if c.title], resp.followups
        except (Exception, asyncio.TimeoutError) as e:
            msg = str(e)
            if i == tries - 1:
                return f"[ERROR] {type(e).__name__}: {msg[:80]}", [], []
            await asyncio.sleep(delay if ("503" in msg or "429" in msg) else 5)
            delay = min(int(delay * 1.5), 90)


async def readiness(rag):
    print("flash-lite 가용성 + 봇5 RAG 인덱싱 확인(503 인내)...", flush=True)
    for attempt in range(13):
        ans, cites, _ = await call(
            rag, 5, "제공된 문서 근거로 한 줄로 답하라.",
            "축복자녀-1세 매칭확정자의 변경된 연령 기준은?")
        if ans.startswith("[ERROR]"):
            print(f"  시도 {attempt+1}: 미가용({ans[:60]}) — 180초 대기", flush=True)
            await asyncio.sleep(180); continue
        if "25" in ans:
            print(f"  시도 {attempt+1}: 가용 + RAG 검색 확인(cites={len(cites)})", flush=True)
            return True
        print(f"  시도 {attempt+1}: 응답하나 RAG 미검색 — 60초 대기", flush=True)
        await asyncio.sleep(60)
    return False


async def main():
    questions = load_questions()
    bots = load_bots()
    print(f"난이도 '{DIFFICULTY}' 문항 {len(questions)}개 · 봇 {len(bots)}개", flush=True)
    for q in questions:
        print(f"  {q['id']} {q['category']} | {q['q'][:30]}", flush=True)

    rag = GeminiRAGService()
    if not await readiness(rag):
        print("  ⚠️ 준비 실패 — 중단(재실행).", flush=True); return

    bots_meta = [{"id": b["id"], "name": b["name"], "llm_model": b["llm_model"],
                  "prompt_len": len(b["system_prompt"] or ""),
                  "성격": BOT_META[b["id"]]["성격"], "rag_docs": BOT_META[b["id"]]["rag_docs"],
                  "slug": FILE_SLUG[b["id"]]}
                 for b in bots]
    responses = []
    for b in bots:
        sp = b["system_prompt"] or ""
        print(f"\n=== 봇 {b['id']} {b['name']} (prompt_len={len(sp)}) ===", flush=True)
        for q in questions:
            ans, cites, fups = await call(rag, b["id"], sp, q["q"])
            responses.append({"bot_id": b["id"], "qid": q["id"], "answer": ans,
                              "citations": cites, "followups": fups})
            flag = " [ERROR]" if ans.startswith("[ERROR]") else ""
            print(f"  {q['id']} ans_len={len(ans):>4} cites={len(cites)}{flag}", flush=True)
            await asyncio.sleep(5)
        # 봇 단위 부분저장(중단 대비)
        OUT.write_text(json.dumps({
            "meta": {"model": MODEL, "temperature": TEMPERATURE, "difficulty": DIFFICULTY,
                     "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                     "note": "정답지는 초안(가정부장 확정 미반영)"},
            "bots": bots_meta, "questions": questions, "responses": responses,
        }, ensure_ascii=False, indent=2), encoding="utf-8")

    err = sum(1 for r in responses if r["answer"].startswith("[ERROR]"))
    print(f"\n저장: {OUT}  (응답 {len(responses)}건, ERROR {err}건)", flush=True)


if __name__ == "__main__":
    asyncio.run(main())
