# 원본 응답 양식(12컬럼) 그대로에 컬럼 3개를 추가해 위험·저점 72건만 필터링한 엑셀 빌드 스크립트
import json
import pathlib

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from _build_report import CATEGORIES  # id→사유 범주 매핑 재사용

OUT_DIR = pathlib.Path(__file__).parent
DATA = OUT_DIR / "_data" / "responses.json"
OUT_XLSX = OUT_DIR / "3주차_레드팀_CD_지정답변_작성목록(원본양식).xlsx"

# (2).xlsx 원본 헤더 그대로 + 추가 3컬럼
ORIG_HEADERS = [
    "타임스탬프",
    "1. 피드백 제출자 정보",
    "2. AI 챗봇에 입력한 질문 (원문)",
    "3. AI 챗봇의 응답 (C)",
    "4. AI 챗봇의 응답 (D)",
    "5. 응답이 적절한 챗봇",
    "6. 응답이 적절한 챗봇에 대한 적절성 및 유용성 평가",
    "7. 응답이 좋았던 챗봇에 대해 좋았던 점",
    "8. 응답이 좋았던 챗봇에 대해 아쉬웠던 점",
    "9. 응답이 좋았던 챗봇에 대해 보완, 제안 할 점",
    "10. 응답이 좋았던 챗봇에 대한 위험도",
    "11. 기타 의견 또는 건의 사항 (예: 챗봇 사용 중 발견한 버그, 시스템 개선 의견 등)",
]
ADD_HEADERS = ["사유 범주", "지정 답변 (직접 작성)", "피드백 반영 여부"]
HEADERS = ORIG_HEADERS + ADD_HEADERS
# 원본 12 + 추가 3 컬럼 너비
WIDTHS = [16, 12, 40, 50, 50, 16, 12, 30, 30, 30, 18, 26, 24, 50, 14]
DV_OPTIONS = "미반영,작성중,반영완료"


def build_id_tags():
    """각 응답 id에 붙일 사유 카테고리 태그를 그룹별로 역매핑."""
    id_cats = {}  # id -> [(group_key, cat_name), ...]
    for gkey, group in CATEGORIES.items():
        for cat in group["cats"]:
            for i in cat["ids"]:
                id_cats.setdefault(i, []).append((gkey, cat["name"]))
    return id_cats


def severity_label(d):
    parts = []
    if d["risk"] == "상":
        parts.append("위험도 상")
    elif d["risk"] == "중":
        parts.append("위험도 중")
    if d.get("score") == 1:
        parts.append("1점")
    elif d.get("score") == 2:
        parts.append("2점")
    return " · ".join(parts)


def category_cell(d, id_cats):
    """추가 컬럼: #id · 위험도/점수 태그 + 사유 범주명(중복 제거, 다중 표기)."""
    sev = severity_label(d)
    names = []
    for _g, name in id_cats.get(d["id"], []):
        if name not in names:
            names.append(name)
    head = f"#{d['id']} · {sev}" if sev else f"#{d['id']}"
    return f"{head}\n{' / '.join(names)}"


def main():
    data = json.loads(DATA.read_text(encoding="utf-8"))
    by_id = {d["id"]: d for d in data}
    id_cats = build_id_tags()

    # 대상 = CATEGORIES 등장 id 합집합(상∪중∪저점), 중복 제거 후 위험도→점수→id 정렬
    target_ids = sorted(id_cats.keys())
    risk_rank = {"상": 0, "중": 1, "하": 2, "없음": 3}
    target_ids.sort(key=lambda i: (
        risk_rank.get(by_id[i]["risk"], 9),
        by_id[i]["score"] if by_id[i]["score"] is not None else 9,
        i,
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "위험·저점 72건"

    # 1행: 제목
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws.cell(row=1, column=1,
                    value="3주차 레드팀 · 라이브 C/D — 원본 응답 양식 + 지정답변 "
                          f"(위험도 상·중 + 적절성 1·2점 {len(target_ids)}건)")
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="left", vertical="center")

    # 2행: 헤더 (원본=남색, 추가=강조색)
    fill_orig = PatternFill("solid", fgColor="1F3864")
    fill_add = PatternFill("solid", fgColor="7C3A12")
    head_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = fill_add if h in ADD_HEADERS else fill_orig
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = WIDTHS[ci - 1]

    # 3행~: 데이터 (원본 12열 그대로 + 추가 3열)
    top_wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for ri, i in enumerate(target_ids, start=3):
        d = by_id[i]
        row = [
            d.get("ts", ""),
            d.get("evaluator", ""),
            d.get("question", ""),
            d.get("respC", ""),
            d.get("respD", ""),
            d.get("prefRaw", ""),
            d.get("score"),
            d.get("good", ""),
            d.get("bad", ""),
            d.get("suggest", ""),
            d.get("riskRaw", ""),
            d.get("etc", ""),
            category_cell(d, id_cats),  # 추가: 사유 범주
            "",                          # 추가: 지정 답변(직접 작성)
            "",                          # 추가: 피드백 반영 여부(드롭다운)
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = center if ci in (7, 15) else top_wrap
            c.border = border

    # 틀고정: 헤더(2행) 고정
    ws.freeze_panes = "A3"

    # 피드백 반영 여부(O열, 15번째) 드롭다운
    last = len(target_ids) + 2
    dv = DataValidation(type="list", formula1=f'"{DV_OPTIONS}"', allow_blank=True)
    dv.add(f"O3:O{last}")
    ws.add_data_validation(dv)

    wb.save(OUT_XLSX)
    print(f"대상 {len(target_ids)}건 · 컬럼 {len(HEADERS)}개(원본 12 + 추가 3) → {OUT_XLSX}")
    print(f"첫 항목 #{target_ids[0]} ({severity_label(by_id[target_ids[0]])})")


if __name__ == "__main__":
    main()
