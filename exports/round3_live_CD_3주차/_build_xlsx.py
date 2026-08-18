# 인쇄본 72개 항목을 지정답변 작성용 엑셀(8열)로 출력하는 빌드 스크립트
import json
import pathlib

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from _build_report import CATEGORIES  # id→사유 범주 매핑 재사용

OUT_DIR = pathlib.Path(__file__).parent
DATA = OUT_DIR / "_data" / "responses.json"
OUT_XLSX = OUT_DIR / "3주차_레드팀_CD_지정답변_작성목록.xlsx"

HEADERS = ["번호", "사유 범주", "질문", "AI 응답 C", "AI 응답 D",
           "이슈 요약", "지정 답변 (직접 작성)", "피드백 반영 여부"]
WIDTHS = [7, 26, 42, 52, 52, 38, 52, 14]
DV_OPTIONS = "미반영,작성중,반영완료"


def build_id_tags():
    """각 응답 id에 붙일 사유 카테고리 태그를 그룹별로 역매핑(_build_print와 동일)."""
    id_cats = {}  # id -> [(group_key, cat_name), ...]
    for gkey, group in CATEGORIES.items():
        for cat in group["cats"]:
            for i in cat["ids"]:
                id_cats.setdefault(i, []).append((gkey, cat["name"]))
    return id_cats


def severity_label(d):
    """위험도/점수 기반 핵심 태그 문자열."""
    parts = []
    if d["risk"] == "상":
        parts.append("위험도 상")
    elif d["risk"] == "중":
        parts.append("위험도 중")
    sc = d.get("score")
    if sc == 1:
        parts.append("1점")
    elif sc == 2:
        parts.append("2점")
    return " · ".join(parts)


def category_cell(d, id_cats):
    """B열: 위험도/점수 태그 + 사유 범주명(중복 제거, 다중 표기)."""
    sev = severity_label(d)
    names = []
    for _gkey, name in id_cats.get(d["id"], []):
        if name not in names:
            names.append(name)
    cats = " / ".join(names)
    return f"{sev}\n{cats}" if sev else cats


def issue_cell(d):
    """F열: 아쉬운 점 + 보완·제안 종합. 둘 다 비면 위험도 사유/기타로 폴백."""
    parts = []
    if d.get("bad"):
        parts.append(f"아쉬운점: {d['bad']}")
    if d.get("suggest"):
        parts.append(f"보완: {d['suggest']}")
    if not parts:
        if d.get("riskRaw"):
            parts.append(f"위험도: {d['riskRaw']}")
        if d.get("etc"):
            parts.append(f"기타: {d['etc']}")
    return "\n".join(parts)


def main():
    data = json.loads(DATA.read_text(encoding="utf-8"))
    by_id = {d["id"]: d for d in data}
    id_cats = build_id_tags()

    # 대상 = CATEGORIES에 등장하는 모든 id의 합집합(상∪중∪저점), 중복 제거
    target_ids = sorted(id_cats.keys())
    risk_rank = {"상": 0, "중": 1, "하": 2, "없음": 3}
    target_ids.sort(key=lambda i: (
        risk_rank.get(by_id[i]["risk"], 9),
        by_id[i]["score"] if by_id[i]["score"] is not None else 9,
        i,
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "지정답변 작성목록"

    # 1행: 제목/안내
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws.cell(row=1, column=1,
                    value="3주차 레드팀 · 라이브 C/D — 지정답변 작성목록 "
                          f"(위험도 상·중 + 적절성 1·2점 {len(target_ids)}건)")
    title.font = Font(bold=True, size=12)
    title.alignment = Alignment(horizontal="left", vertical="center")

    # 2행: 헤더
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = w

    # 3행~: 데이터
    top_wrap = Alignment(vertical="top", wrap_text=True)
    center = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for ri, i in enumerate(target_ids, start=3):
        d = by_id[i]
        row = [
            f"#{d['id']}",
            category_cell(d, id_cats),
            d.get("question", ""),
            d.get("respC", ""),
            d.get("respD", ""),
            issue_cell(d),
            "",   # G 지정 답변(직접 작성) — 빈칸
            "",   # H 피드백 반영 여부 — 빈칸(드롭다운)
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = center if ci in (1, 8) else top_wrap
            c.border = border

    # 틀고정: 헤더(2행)+번호열(A) 고정 → B3
    ws.freeze_panes = "B3"

    # H열 드롭다운(미반영/작성중/반영완료)
    last = len(target_ids) + 2
    dv = DataValidation(type="list", formula1=f'"{DV_OPTIONS}"', allow_blank=True)
    dv.add(f"H3:H{last}")
    ws.add_data_validation(dv)

    wb.save(OUT_XLSX)
    print(f"대상 {len(target_ids)}건 → {OUT_XLSX}")
    print(f"첫 항목 #{target_ids[0]} ({severity_label(by_id[target_ids[0]])})")


if __name__ == "__main__":
    main()
