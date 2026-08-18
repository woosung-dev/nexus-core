# FAQ 자동응답 발동 건을 제외했을 때 1주차 vs 3주차 적절성 점수 변화를 비교하는 HTML 리포트 빌더
"""
1주차(레드팀 응답 (6).xlsx, 단일응답 적절성 점수) vs
3주차(라이브 C/D 블라인드 responses.json, 적절성 점수)를 비교한다.

FAQ 자동응답 발동 = 응답 원문이 FAQ 지정답변과 (공백정규화 후) 거의 일치(ratio>=0.92).
C·D가 같은 FAQ·임계값·임베딩을 쓰므로 단독 발동은 0건 → '하나라도/둘 다'가 동일.
따라서 단일 'FAQ 제외' 칼럼으로 비교한다.

산출물:
  - faq_tagged.json  : 3주차 268건 + faqC/faqD/faqNo 태그
  - 1주차_vs_3주차_FAQ제외_비교.html
"""
import json
import re
import difflib
from pathlib import Path
from collections import Counter

import openpyxl

ROOT = Path(__file__).resolve().parent
DOWNLOADS = Path("/Users/woosung/Downloads")
FAQ_XLSX = DOWNLOADS / "블레싱네비게이션_FAQ_지정답변_작성완료.xlsx"
WEEK1_XLSX = DOWNLOADS / "축복·가정관리 AI 상담 챗봇 테스트 및 피드백 (레드팀)(응답) (6).xlsx"
WEEK3_JSON = ROOT / "_data" / "responses.json"
MATCH_TH = 0.92


def norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def load_faqs():
    wb = openpyxl.load_workbook(FAQ_XLSX, data_only=True)
    ws = wb["FAQ 전체 목록"]
    faqs = []
    for ri in range(3, ws.max_row + 1):
        no = ws.cell(ri, 1).value
        q = ws.cell(ri, 3).value
        ans = ws.cell(ri, 8).value
        if ans and str(ans).strip():
            faqs.append({"no": int(no), "q": norm(q), "ans": norm(ans)})
    return faqs


def best_faq(text, faqs):
    t = norm(text)
    if not t:
        return None, 0.0
    b = max(faqs, key=lambda f: difflib.SequenceMatcher(None, f["ans"], t).ratio())
    return b, difflib.SequenceMatcher(None, b["ans"], t).ratio()


def load_week1_scores():
    wb = openpyxl.load_workbook(WEEK1_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    scores = []
    for ri in range(2, ws.max_row + 1):
        v = ws.cell(ri, 6).value
        if v is None or str(v).strip() == "":
            continue
        m = re.search(r"([1-5])", str(v))
        if m:
            scores.append(int(m.group(1)))
    return scores


def stats(scores):
    n = len(scores)
    avg = sum(scores) / n if n else 0.0
    dist = {s: 0 for s in range(1, 6)}
    for s in scores:
        dist[int(round(s))] += 1
    low = sum(1 for s in scores if s <= 2)
    high = sum(1 for s in scores if s >= 4)
    return {
        "n": n,
        "avg": avg,
        "dist": dist,
        "low": low,
        "low_pct": low / n * 100 if n else 0,
        "high": high,
        "high_pct": high / n * 100 if n else 0,
    }


def main():
    faqs = load_faqs()
    week3 = json.load(open(WEEK3_JSON))

    tagged = []
    fired = []
    for r in week3:
        bc, rc = best_faq(r.get("respC"), faqs)
        bd, rd = best_faq(r.get("respD"), faqs)
        fc, fd = rc >= MATCH_TH, rd >= MATCH_TH
        row = dict(r)
        row["faqC"] = fc
        row["faqD"] = fd
        row["faqNoC"] = bc["no"] if fc else None
        row["faqNoD"] = bd["no"] if fd else None
        tagged.append(row)
        if fc or fd:
            fired.append(row)

    (ROOT / "_data" / "faq_tagged.json").write_text(
        json.dumps(tagged, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    w1 = load_week1_scores()
    w3_all = [r["score"] for r in week3 if isinstance(r.get("score"), (int, float))]
    fired_ids = {r["id"] for r in fired}
    w3_excl = [
        r["score"]
        for r in week3
        if isinstance(r.get("score"), (int, float)) and r["id"] not in fired_ids
    ]

    s_w1 = stats(w1)
    s_w3_all = stats(w3_all)
    s_w3_excl = stats(w3_excl)
    s_fired = stats([r["score"] for r in fired if isinstance(r.get("score"), (int, float))])

    only_one = sum(1 for r in fired if r["faqC"] != r["faqD"])

    html = build_html(
        faqs, fired, s_w1, s_w3_all, s_w3_excl, s_fired, only_one
    )
    out = ROOT / "1주차_vs_3주차_FAQ제외_비교.html"
    out.write_text(html, encoding="utf-8")

    print(f"FAQ 발동 {len(fired)}건 (단독 발동 {only_one}건)")
    print(f"1주차       n={s_w1['n']} avg={s_w1['avg']:.3f}")
    print(f"3주차 전체   n={s_w3_all['n']} avg={s_w3_all['avg']:.3f}")
    print(f"3주차 FAQ제외 n={s_w3_excl['n']} avg={s_w3_excl['avg']:.3f}")
    print(f"발동 14건 평균 {s_fired['avg']:.3f}")
    print(f"-> {out}")


def bar(dist, total, color):
    """1~5 점수 분포 막대 (가로 누적)"""
    seg = []
    palette = {1: "#B5321E", 2: "#D98324", 3: "#C9A227", 4: "#5C8A7A", 5: "#1E3A34"}
    for s in range(1, 6):
        pct = dist[s] / total * 100 if total else 0
        if pct <= 0:
            continue
        seg.append(
            f'<span class="seg" style="width:{pct:.1f}%;background:{palette[s]}" '
            f'title="{s}점 {dist[s]}건 ({pct:.1f}%)">{s}</span>'
        )
    return f'<div class="distbar">{"".join(seg)}</div>'


def build_html(faqs, fired, s_w1, s_w3_all, s_w3_excl, s_fired, only_one):
    faq_q = {f["no"]: f["q"] for f in faqs}
    d_all = s_w3_all["avg"] - s_w1["avg"]
    d_excl = s_w3_excl["avg"] - s_w1["avg"]
    lift = s_w3_excl["avg"] - s_w3_all["avg"]

    fired_sorted = sorted(fired, key=lambda r: (r["score"], r["id"]))
    fired_rows = ""
    for r in fired_sorted:
        no = r["faqNoC"] or r["faqNoD"]
        scls = "sc1" if r["score"] <= 2 else ("sc3" if r["score"] == 3 else "sc4")
        fired_rows += f"""
        <tr>
          <td class="mono">{r['id']}</td>
          <td>{esc(r['question'])}</td>
          <td class="mono center {scls}">{r['score']:.0f}</td>
          <td class="center"><span class="faqtag">FAQ #{no}</span></td>
          <td>{esc(faq_q.get(no, ''))}</td>
          <td>{esc(r.get('evaluator',''))}</td>
        </tr>"""

    def card(title, st, sub=""):
        return f"""
        <div class="card">
          <div class="card-h">{title}</div>
          <div class="bignum">{st['avg']:.3f}<span class="unit"> / 5</span></div>
          <div class="card-sub">n = {st['n']}{sub}</div>
          {bar(st['dist'], st['n'], '')}
          <div class="card-meta">낮음(1·2점) {st['low']}건 · {st['low_pct']:.1f}% &nbsp;|&nbsp; 높음(4·5점) {st['high']}건 · {st['high_pct']:.1f}%</div>
        </div>"""

    return f"""<!doctype html>
<html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>1주차 vs 3주차 · FAQ 자동응답 제외 점수 비교</title>
<style>
  :root {{
    --paper:#F1EDE2; --ink:#17150F; --pine:#1E3A34; --brass:#B08524;
    --garnet:#B5321E; --sage:#5C8A7A; --line:#d9d2c2; --card:#fbf9f3;
  }}
  * {{ box-sizing:border-box; }}
  body {{ margin:0; background:var(--paper); color:var(--ink);
    font-family:"Noto Sans KR","Apple SD Gothic Neo",system-ui,sans-serif; line-height:1.6; }}
  .wrap {{ max-width:1100px; margin:0 auto; padding:48px 28px 80px; }}
  h1 {{ font-family:"Noto Serif KR",serif; font-size:30px; margin:0 0 6px; letter-spacing:-.5px; }}
  .lead {{ color:#5c574a; margin:0 0 4px; }}
  .mono {{ font-family:"IBM Plex Mono","SF Mono",monospace; font-variant-numeric:tabular-nums; }}
  .secno {{ color:var(--garnet); font-weight:700; margin-right:8px; }}
  h2 {{ font-family:"Noto Serif KR",serif; font-size:21px; margin:44px 0 14px;
    padding-bottom:8px; border-bottom:2px solid var(--pine); }}
  .cards {{ display:grid; grid-template-columns:repeat(3,1fr); gap:16px; margin:18px 0; }}
  .card {{ background:var(--card); border:1px solid var(--line); border-radius:12px; padding:18px 18px 16px;
    box-shadow:0 1px 0 #fff inset, 0 1px 3px rgba(0,0,0,.04); }}
  .card-h {{ font-weight:700; font-size:14px; color:var(--pine); margin-bottom:6px; }}
  .bignum {{ font-family:"IBM Plex Mono",monospace; font-size:38px; font-weight:600; line-height:1; }}
  .unit {{ font-size:15px; color:#8a8472; font-weight:400; }}
  .card-sub {{ color:#6b6557; font-size:13px; margin:6px 0 12px; }}
  .card-meta {{ font-size:12px; color:#6b6557; margin-top:10px; }}
  .distbar {{ display:flex; height:22px; border-radius:5px; overflow:hidden; border:1px solid var(--line); }}
  .seg {{ color:#fff; font-size:11px; font-weight:700; display:flex; align-items:center; justify-content:center;
    min-width:14px; }}
  .delta {{ display:inline-block; padding:2px 9px; border-radius:20px; font-weight:700; font-size:13px;
    font-family:"IBM Plex Mono",monospace; }}
  .up {{ background:#e3efe9; color:var(--pine); }}
  .down {{ background:#f6e0dc; color:var(--garnet); }}
  table {{ width:100%; border-collapse:collapse; font-size:13.5px; background:var(--card);
    border:1px solid var(--line); border-radius:10px; overflow:hidden; }}
  th,td {{ padding:9px 11px; text-align:left; border-bottom:1px solid var(--line); vertical-align:top; }}
  th {{ background:#efe9da; font-size:12px; color:#5c574a; font-weight:700; }}
  td.center,th.center {{ text-align:center; }}
  tr:last-child td {{ border-bottom:none; }}
  .faqtag {{ background:var(--garnet); color:#fff; padding:2px 8px; border-radius:5px; font-size:11.5px;
    font-weight:700; white-space:nowrap; font-family:"IBM Plex Mono",monospace; }}
  .sc1 {{ color:var(--garnet); font-weight:700; }}
  .sc3 {{ color:var(--brass); font-weight:700; }}
  .sc4 {{ color:var(--pine); font-weight:700; }}
  .summary {{ background:var(--pine); color:#f1ede2; border-radius:12px; padding:22px 24px; margin:18px 0; }}
  .summary b {{ color:#fff; }}
  .summary .row {{ display:flex; justify-content:space-between; padding:7px 0; border-bottom:1px solid rgba(255,255,255,.14);
    font-size:14.5px; }}
  .summary .row:last-child {{ border-bottom:none; }}
  .note {{ background:#fbf4e3; border-left:4px solid var(--brass); padding:12px 16px; border-radius:0 8px 8px 0;
    font-size:13.5px; color:#5c540e; margin:14px 0; }}
  .foot {{ color:#8a8472; font-size:12px; margin-top:40px; border-top:1px solid var(--line); padding-top:14px; }}
  @media (max-width:760px) {{ .cards {{ grid-template-columns:1fr; }} }}
</style></head>
<body><div class="wrap">

  <h1>FAQ 자동응답 제외 시 1주차 → 3주차 점수 비교</h1>
  <p class="lead">3주차 라이브 C·D 블라인드 테스트에서 FAQ 자동응답(지정답변 그대로 출력)이 발동한 {len(fired)}건을 제외했을 때, 1주차 대비 적절성 점수가 어떻게 달라지는지 비교한다.</p>
  <p class="lead mono" style="font-size:12.5px">적절성 평가(1~5점) · 1주차=단일응답 설문 · 3주차=C/D 블라인드 1문항당 1점 · FAQ 발동 판정=응답 원문이 FAQ 지정답변과 ≥{int(MATCH_TH*100)}% 일치</p>

  <div class="summary">
    <div class="row"><span>1주차 평균 (FAQ 없음 · 베이스라인)</span><b class="mono">{s_w1['avg']:.3f}</b></div>
    <div class="row"><span>3주차 전체 평균</span><b class="mono">{s_w3_all['avg']:.3f} &nbsp;(<span class="delta {'up' if d_all>=0 else 'down'}">{d_all:+.3f}</span> vs 1주차)</b></div>
    <div class="row"><span>3주차 · FAQ 발동 {len(fired)}건 제외</span><b class="mono">{s_w3_excl['avg']:.3f} &nbsp;(<span class="delta {'up' if d_excl>=0 else 'down'}">{d_excl:+.3f}</span> vs 1주차)</b></div>
    <div class="row"><span>FAQ 제외로 인한 3주차 점수 보정</span><b class="mono"><span class="delta {'up' if lift>=0 else 'down'}">{lift:+.3f}</span></b></div>
  </div>

  <div class="note">
    <b>'하나라도 걸림' = '둘 다 걸림' (단독 발동 {only_one}건).</b>
    C(id6)·D(id7) 라이브 봇은 동일한 FAQ 세트·임계값·임베딩 모델을 사용하므로, 한 질문이 FAQ에 매칭되면 두 봇이 함께 발동한다.
    실제로 발동 {len(fired)}건 전부 C·D가 같은 FAQ로 발동했고 단독 발동은 0건이라, 두 시나리오 결과가 같다. 그래서 단일 'FAQ 제외' 칼럼으로 통합했다.
  </div>

  <h2><span class="secno">1</span>점수 카드 비교</h2>
  <div class="cards">
    {card('1주차 (베이스라인)', s_w1, ' · FAQ 미존재')}
    {card('3주차 전체', s_w3_all, ' · FAQ 발동 포함')}
    {card('3주차 · FAQ 제외', s_w3_excl, f' · {len(fired)}건 제외')}
  </div>
  <div class="note" style="border-color:var(--garnet);color:#7a2417;background:#f9ece9">
    FAQ 발동 {len(fired)}건의 평균은 <b class="mono">{s_fired['avg']:.3f}점</b>으로 3주차 전체 평균({s_w3_all['avg']:.3f})보다 크게 낮다.
    이 저득점 묶음을 제외하니 3주차 평균이 {s_w3_all['avg']:.3f} → {s_w3_excl['avg']:.3f} ({lift:+.3f})로 올라가고,
    1주차 대비 상승폭이 {d_all:+.3f} → {d_excl:+.3f}로 더 뚜렷해진다.
  </div>

  <h2><span class="secno">2</span>FAQ 자동응답 발동 {len(fired)}건 (TAG 대상)</h2>
  <table>
    <thead><tr>
      <th>ID</th><th>질문 (3주차)</th><th class="center">점수</th><th class="center">TAG</th>
      <th>매칭된 FAQ 질문</th><th>평가자</th>
    </tr></thead>
    <tbody>{fired_rows}
    </tbody>
  </table>
  <div class="note">
    대표 오발동 — "축복을 왜 받아야 해요?"(ID 21·258)가 <b>FAQ #18 가정회비</b> 지정답변에 임베딩 오매칭되어,
    C·D가 글자 단위로 동일한 가정회비 안내문을 출력했다(점수 1.0). 질문 의도와 무관한 답이 나가 저득점을 받은 전형적 사례다.
  </div>

  <div class="foot">
    소스 — 3주차: exports/round3_live_CD_3주차/_data/responses.json · 1주차: Downloads/…(레드팀)(응답) (6).xlsx (적절성 col) · FAQ 지정답변: Downloads/블레싱네비게이션_FAQ_지정답변_작성완료.xlsx<br>
    재생성 — <span class="mono">backend/.venv/bin/python exports/round3_live_CD_3주차/_build_faq_compare.py</span> ·
    주의 — 1주차(단일응답 설문)와 3주차(C/D 블라인드)는 평가 설계가 달라 집단 평균 수준의 추이 비교로만 해석한다.
  </div>

</div></body></html>"""


def esc(s):
    return (
        str(s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


if __name__ == "__main__":
    main()
