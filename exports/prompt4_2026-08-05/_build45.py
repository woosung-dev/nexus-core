# 45개 선정본(2026-08-05 수정본 xlsx) + 불변제약 10 → exports/regression/questions.json
#
# 기존 50문항 세트(A위험상20+B위험중20+C10)는 폐기하고 관리자 선정 45개로 갈아탄다
# (docs/architecture/handoff-qa-golden-45set-2026-08-05.md §0). C 불변제약 10건은
# _build_questions.py 의 정의를 **그대로 재사용**한다 — 골든이 있어 L3 채점이 성립하고
# 7/29·8/03 리포트와 같은 축(Critical 건수·할루시율)으로 세로 비교가 된다.
#
# 관리자 키워드는 `anchors` 로 넣고 `must_any` 에는 넣지 않는다.
#   이유: "축복의 기본 의의" 같은 개념구는 _l2.py 의 부분문자열 판정으로 못 잡아
#   거짓실패가 쏟아진다. 의미 판정은 _anchor.py(codex) 가 한다.
import asyncio
import importlib.util
import json
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path("/Users/woosung/project/agy-project/nexus-core")
sys.path.insert(0, str(ROOT / "backend"))

from sqlalchemy import text  # noqa: E402

from app.core.database import async_session  # noqa: E402

XLSX = Path.home() / "Downloads" / "블레싱네비게이션_3주차_핵심질문_45개_규정집키워드.xlsx"
REG = ROOT / "exports" / "regression"
OUT = REG / "questions.json"
BACKUP = REG / "questions_50_2026-08-04.json"

# C_ITEMS 를 복제하지 않고 원본 모듈에서 가져온다 (파일명이 숫자로 시작하지 않아 import 가능)
_spec = importlib.util.spec_from_file_location("_bq", REG / "_build_questions.py")
_bq = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_bq)
C_ITEMS = _bq.C_ITEMS


def nfc(s):
    return unicodedata.normalize("NFC", str(s or "")).strip()


def norm(s):
    """DB question_norm 과 같은 규칙 — NFC + 공백·구두점 제거."""
    t = nfc(s)
    return "".join(ch for ch in t if ch.isalnum())


def read_xlsx():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    ws = wb["선정질문_45개"]
    hdr = [nfc(c) for c in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
    sel = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=5, values_only=True)
           if r[0] is not None]

    ws = wb["답변키워드_45개"]
    kw = {}
    for r in ws.iter_rows(min_row=5, values_only=True):
        if r[0] is None:
            continue
        kw[int(r[0])] = {"cat": nfc(r[1]), "q": nfc(r[2]),
                         "keywords": nfc(r[3]), "status": nfc(r[4])}

    assert len(sel) == 45 and len(kw) == 45, f"45 아님: sel={len(sel)} kw={len(kw)}"
    return sel, kw


def split_keywords(s):
    """'A, B·C, D' → ['A','B','C','D'] — 관리자가 쉼표와 가운뎃점을 섞어 쓴다."""
    out = []
    for part in nfc(s).replace("·", ",").split(","):
        t = part.strip()
        if t:
            out.append(t)
    return out


async def main():
    sel, kw = read_xlsx()

    async with async_session() as s:
        rows = (await s.execute(text(
            "SELECT id, question, question_norm, risk, category FROM redteam_question_groups"
        ))).mappings().all()

    by_norm = {}
    for r in rows:
        by_norm.setdefault(norm(r["question"]), r)
        if r["question_norm"]:
            by_norm.setdefault(norm(r["question_norm"]), r)

    picked, missed = [], []
    for rec in sel:
        no = int(rec["번호"])
        q = nfc(rec["질문 원문"])
        k = kw[no]
        assert norm(k["q"]) == norm(q), f"#{no} 두 탭의 질문 원문 불일치"

        hit = by_norm.get(norm(q))
        if hit is None:
            missed.append((no, q[:40]))

        risk = nfc(rec["분석 위험도"])
        picked.append({
            "bucket": "A" if risk == "상" else "B",
            "gid": hit["id"] if hit is not None else None,
            "no": no,
            "norm": norm(q),
            "q": q,
            "risk": risk,
            "cat": nfc(rec["카테고리"]),
            "subtype": nfc(rec["세부유형"]),
            "difficulty": nfc(rec["난이도"]),
            "risk_original": nfc(rec["원본 위험도"]),
            "submitter": nfc(rec["제출자"]),
            # 관리자가 붙인 것 — 이번 채점의 기준
            "anchors": split_keywords(k["keywords"]),
            "evidence_status": k["status"],
            # 기계 판정(_l2)용은 비워 둔다. 의미 판정은 _anchor.py 가 한다.
            "must_any": [],
            "must_not": [],
        })

    for i, c in enumerate(C_ITEMS, 1):
        picked.append({"bucket": "C", "gid": None, "no": None, "norm": None, "q": c["q"],
                       "risk": None, "cat": f"불변제약:{c['area']}",
                       "rubric": c["rubric"], "gate": c["gate"], "golden": c["golden"],
                       "must_any": c["must_any"], "must_not": c["must_not"],
                       "fail": c["fail"], "cid": f"C{i:02d}"})

    counts = {b: sum(p["bucket"] == b for p in picked) for b in "ABC"}
    counts["total"] = len(picked)

    if OUT.exists() and not BACKUP.exists():
        BACKUP.write_text(OUT.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"기존 50문항 세트 백업 → {BACKUP.name}")

    OUT.write_text(json.dumps({
        "version": "regression-45set-2026-08-05",
        "source": {"A": "관리자 선정 45개 중 분석 위험도 '상'",
                   "B": "관리자 선정 45개 중 그 외",
                   "C": "평가셋_루브릭_round3.md 불변제약 10 (_build_questions.py 재사용)",
                   "xlsx": XLSX.name},
        "counts": counts, "items": picked}, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"A {counts['A']} · B {counts['B']} · C {counts['C']} = {counts['total']}건 → {OUT.name}")
    print(f"gid 매칭 {45 - len(missed)}/45")
    for no, q in missed:
        print(f"  ⚠ 미매칭 #{no} {q}")
    st = {}
    for p in picked:
        if p["bucket"] != "C":
            st[p["evidence_status"]] = st.get(p["evidence_status"], 0) + 1
    print(f"규정집 근거 상태: {st}")
    print(f"앵커 총 {sum(len(p.get('anchors', [])) for p in picked)}개 "
          f"(문항 평균 {sum(len(p.get('anchors', [])) for p in picked)/45:.1f})")


if __name__ == "__main__":
    asyncio.run(main())
