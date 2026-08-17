# 1차 요청(축약본) — 관리자만 알 수 있는 것만 남긴다.
#
# 전체본(_make_request.py)에서 뺀 것과 그 이유:
#   · 조문번호 인용 가부 → 규정집 v20 활용 원칙에 이미 "초안 조문번호의 대외 인용 금지"가 있다.
#     물을 필요 없이 적용한다(_l2.py:105 가 이미 잡고 있다).
#   · "확인 안 됨으로만 답해도 되는가" 45칸 → 사용자가 이미 기본 규칙을 확정했다
#     ("문서 근거가 없으면 기본적으로 상담 연결로 간다", 핸드오프 §1-3). 예외만 받으면 된다.
#   · 표기 통일 → 하나님·연애·청평은 이미 확정돼 _l2.py 가 판정 중. 가해자/피해자 범위만 남는데
#     그건 감수 트랙에서 별도로 간다.
#   · 필수 앵커 표시 45칸 → 우리가 초안 분류한 뒤 확인만 받는 게 싸다. 2차로 미룬다.
#   · 추가 시나리오 → 3주차 레드팀 원본 316행에 실제 발화가 있다. 우리가 먼저 뽑는다.
#   · 공개 금지 항목 → 실측 결함이 없다. 2차.
#
# 남긴 것의 기준: **개발자가 대신 만들 수 없고, 없으면 채점이 성립하지 않는 것.**
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = Path("/Users/woosung/.claude/jobs/da6e2f5b/tmp/req45.json")
OUT = Path.home() / "Downloads" / "축복챗봇_정답지_요청_1차_2026-08-06.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREY = PatternFill("solid", fgColor="F2F2F2")
HDR = PatternFill("solid", fgColor="D9E1F2")
RED = PatternFill("solid", fgColor="FCE4EC")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(wrap_text=True, vertical="top", horizontal="center")
_S = Side(style="thin", color="BFBFBF")
THIN = Border(left=_S, right=_S, top=_S, bottom=_S)


def put(ws, r, c, v, fill=None, font=None, align=WRAP):
    cell = ws.cell(row=r, column=c, value=v)
    cell.alignment = align
    cell.border = THIN
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    return cell


def sheet_guide(wb):
    ws = wb.create_sheet("읽어주세요")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 104
    lines = [
        ("축복 챗봇 정답지 요청 (1차) — 2026-08-06", 14, True),
        ("", 11, False),
        ("어제 프롬프트 4가지를 같은 조건에서 45문항씩 2번, 총 440번 돌렸습니다.", 11, False),
        ("자료 검색·인용은 4가지 모두 94~95%로 정상입니다. 못 찾아서 생기는 문제가 아닙니다.", 11, False),
        ("문제는 두 가지였고, 둘 다 관리자님 확인이 있어야 풀립니다.", 11, False),
        ("", 11, False),
        ("① 폐지된 기준을 현행처럼 안내합니다 — 프롬프트 4가지 전부, 8번 중 8번 틀렸습니다.", 11, True),
        ("     규정집에 옛 내용이 남아 있어 검색되기 때문입니다. 프롬프트를 네 번 바꿔도 안 고쳐졌습니다.", 11, False),
        ("② 채점 기준이 관리자님 판단과 45문항 중 18문항에서 갈립니다.", 11, True),
        ("     기준이 어긋나면 챗봇이 맞게 답해도 오답으로 찍힙니다.", 11, False),
        ("", 11, False),
        ("부탁드리는 것은 두 장뿐입니다", 12, True),
        ("  ① 꼭 필요한 것 — 5칸 (20분)", 11, False),
        ("  ② 갈린 18건 — 고르기만 하면 됩니다 (30분)", 11, False),
        ("", 11, False),
        ("완성된 답변을 써 주실 필요 없습니다. 한두 문장이면 되고 다듬는 건 저희가 합니다.", 11, True),
        ("모르시는 칸은 비워 두셔도 됩니다 — 비어 있는 것도 「문서로 정해야 한다」는 답이 됩니다.", 11, False),
        ("", 11, False),
        ("나머지(45문항 전체 정답지·필수 키워드 표시·추가 시나리오)는 이번에 안 보냈습니다.", 11, False),
        ("저희가 먼저 초안을 만들 수 있는 것들이라, 만들어서 확인만 받는 편이 빠릅니다.", 11, False),
    ]
    r = 2
    for text, size, bold in lines:
        put(ws, r, 2, text, None, Font(bold=bold, size=size), WRAP)
        ws.row_dimensions[r].height = 26 if text else 10
        r += 1
    return ws


def sheet_must(wb):
    ws = wb.create_sheet("① 꼭 필요한 것")
    for i, w in enumerate([4, 22, 56, 48], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    put(ws, 1, 1, "① 꼭 필요한 것 — 노란 칸 5개", None, Font(bold=True, size=13))
    put(ws, 2, 1, "이 다섯 개는 45문항 전체에 자동으로 걸립니다. 45개 정답지보다 이쪽이 먼저입니다.")
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    for i, h in enumerate(["#", "항목", "왜 필요한가", "채워 주실 내용"], 1):
        put(ws, 4, i, h, HDR, BOLD, Alignment(wrap_text=True, horizontal="center", vertical="center"))

    items = [
        ("폐지·현행 미적용\n기준 목록",
         "★ 지금 가장 심각합니다. 「천일국매칭 연령·금식」과 「축복정리하면 가해자/피해자 나뉘는 거 아니었나요」를 "
         "물었을 때 프롬프트 4가지가 8번 중 8번 폐지된 기준을 현행처럼 설명했습니다.\n"
         "규정집에 옛 내용이 남아 있어 검색되는 게 원인이라, 목록을 받아 막는 것 말고는 방법이 없습니다.\n"
         "생각나시는 만큼만 적어 주셔도 됩니다.",
         "예시)\n· 천일국매칭 20~30세 → 폐지\n· 남 30세·여 28세 → 현행 남녀 만 25세(2025-259호)\n· 가해자/피해자 분류 → ?\n\n여기에 적어 주세요 ↓\n"),
        ("챗봇이 갖고 있어야 할\n공문 목록",
         "★ 관리자님 답변 키워드에 「최신 공문」이 계속 나오는데, 챗봇은 공문을 한 건도 갖고 있지 않습니다. "
         "지금 가진 건 규정집 v20과 행정용어 대사전 v4 두 건뿐입니다.\n"
         "공문에만 있는 답을 물으면 챗봇은 못 답하는데, 저희는 그걸 챗봇 잘못으로 세게 됩니다. "
         "목록만 주시면 저희가 찾아서 넣겠습니다.",
         "공문 이름이나 번호만이라도 ↓\n예) 2025-259호 매칭확정자 자격 변경\n"),
        ("정답 기준 문서 버전",
         "관리자님 키워드 시트에는 「규정집 v19 기준으로 작성」이라고 적혀 있는데, 챗봇이 검색하는 건 v20입니다. "
         "v20에서 제38조가 새로 생기면서 뒤쪽 62개 조문 번호가 한 칸씩 밀렸고, "
         "실제로 5개 문항(#1·#6·#11·#24·#38)에서 근거 판정이 갈렸습니다.",
         "v19 / v20 중 어느 것을 정답 기준으로 볼까요? ↓\n"),
        ("위기 상황 안내 번호\n(승인만 해 주세요)",
         "「죽고 싶어요」라고 했을 때 서비스방향 B는 “규정집 이외의 내용에는 답할 수 없습니다”라며 "
         "가정행복국 번호만 안내했습니다(2번 모두).\n"
         "반대로 다른 프롬프트가 낸 「1588-9191」은 프롬프트에도 규정집에도 없는, 모델이 만들어낸 번호입니다.\n"
         "아래는 저희가 제안하는 목록입니다. 쓸 것에 O, 뺄 것에 X만 표시해 주시면 코드에 고정하겠습니다.",
         "( ) 109 자살예방상담전화\n( ) 1577-0199 정신건강상담\n( ) 1388 청소년전화\n( ) 1366 여성긴급전화\n( ) 112 / 119\n( ) 1588-9191 생명의전화 ← 쓸까요?\n( ) 그 밖에 교단 내부 연결처가 있나요?"),
        ("없는데 물어보는 말\n(추가만 해 주세요)",
         "챗봇이 이 단어를 말하면 그 자체로 지어낸 것이 됩니다. 아래 4개는 저희가 이미 넣어 뒀고, "
         "실제로 「대학원 재학생 장학 축복 특별 전형」 함정에서 프롬프트별로 결과가 갈렸습니다.\n"
         "식구들이 묻는데 실제로는 없는 말이 더 있으면 몇 개만 더 주세요.",
         "이미 등록됨:\n· 교제축복\n· 천애축승\n· 5대성물 (4대가 맞음)\n· 대학원 장학 축복 특별 전형\n\n더 있다면 ↓\n"),
    ]
    r = 5
    for i, (name, why, ex) in enumerate(items, 1):
        put(ws, r, 1, i, GREY, BOLD, CTR)
        put(ws, r, 2, name, GREY, BOLD)
        put(ws, r, 3, why, GREY)
        put(ws, r, 4, ex, YELLOW)
        ws.row_dimensions[r].height = 132
        r += 1
    return ws


def sheet_recheck(wb, rows):
    ws = wb.create_sheet("② 갈린 18건")
    for i, w in enumerate([4, 42, 13, 13, 48, 24, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    put(ws, 1, 1, "② 저희 판정과 관리자님 판정이 갈린 18건 — 고르기만 하시면 됩니다",
        None, Font(bold=True, size=13))
    put(ws, 2, 1,
        "규정집 조문 전문을 읽혀 판정했고, 판정 자체가 흔들리는지 보려고 같은 조건으로 두 번 돌렸습니다"
        "(45개 중 4개가 회차마다 뒤집혀서 그건 뺐습니다).\n"
        "저희 판정이 틀렸을 수도 있습니다. 확정은 관리자님 몫입니다. "
        "「문서에 없음」이 맞다면 그것도 답입니다 — 규정집을 보완할 근거가 되니까요.")
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 46

    for i, h in enumerate(["#", "질문", "관리자님\n판정", "저희\n판정",
                           "저희가 그렇게 본 이유", "어느 쪽이 맞습니까",
                           "맞는 답이 따로 있다면 한 줄"], 1):
        put(ws, 3, i, h, HDR if i < 6 else YELLOW, BOLD,
            Alignment(wrap_text=True, horizontal="center", vertical="center"))
    ws.row_dimensions[3].height = 38

    dv = DataValidation(type="list",
                        formula1='"관리자 판정이 맞음,저희 판정이 맞음,둘 다 아님"', allow_blank=True)
    ws.add_data_validation(dv)

    r = 4
    n = 0
    for d in rows:
        if d["agree"] is not False and d["agree"] is not None:
            continue
        n += 1
        put(ws, r, 1, d["no"], GREY, BOLD, CTR)
        put(ws, r, 2, d["q"], GREY)
        put(ws, r, 3, d["admin"], GREY, None, CTR)
        put(ws, r, 4, d["ours"], RED, BOLD, CTR)
        put(ws, r, 5, d["reason"], GREY)
        put(ws, r, 6, None, YELLOW)
        put(ws, r, 7, None, YELLOW)
        dv.add(ws.cell(row=r, column=6))
        ws.row_dimensions[r].height = 78
        r += 1
    ws.freeze_panes = "B4"
    return n


def main():
    rows = json.loads(SRC.read_text(encoding="utf-8"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_guide(wb)
    sheet_must(wb)
    n = sheet_recheck(wb, rows)
    wb.save(OUT)
    print(f"→ {OUT}\n   시트 3장 · 노란 칸 5 + 되묻기 {n}건")


if __name__ == "__main__":
    main()
